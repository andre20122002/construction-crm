from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, DecimalField, Count, Q, Case, When
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from datetime import timedelta
from django.utils import timezone
from django.utils.timezone import make_naive
import json
import openpyxl
from decimal import Decimal
from ..models import Transaction, Order, OrderItem, Warehouse, Material, Supplier, AuditLog
from ..forms import PeriodReportForm
from .utils import get_user_warehouses, get_warehouse_balance
from ..decorators import group_required

@login_required
def reports_dashboard(request):
    if not request.user.is_staff: return redirect('index')
    
    warehouse_stats = Transaction.objects.filter(transaction_type='OUT').values('warehouse__name').annotate(
        total_spent=Sum(F('quantity') * F('price'))
    ).order_by('-total_spent')
    
    wh_labels = [item['warehouse__name'] for item in warehouse_stats]
    wh_data = [round(float(item['total_spent'] or 0), 2) for item in warehouse_stats]
    
    top_materials_qs = Transaction.objects.filter(transaction_type='OUT').values('material__name', 'material__unit').annotate(
        total_spent=Sum(F('quantity') * F('price')), total_qty=Sum('quantity')
    ).order_by('-total_spent')[:5]
    
    monthly_stats = Transaction.objects.filter(transaction_type='OUT').annotate(month=TruncMonth('created_at')).values('month').annotate(total=Sum(F('quantity') * F('price'))).order_by('month')
    month_labels = []; month_data = []
    for item in monthly_stats:
        if item['month']:
            month_labels.append(item['month'].strftime('%Y-%m'))
            month_data.append(round(float(item['total'] or 0), 2))
            
    return render(request, 'warehouse/reports.html', {
        'wh_labels': json.dumps(wh_labels), 'wh_data': json.dumps(wh_data),
        'top_materials': top_materials_qs, 'month_labels': json.dumps(month_labels), 'month_data': json.dumps(month_data)
    })

@login_required
def period_report(request):
    today = timezone.now().date()
    form = PeriodReportForm(request.GET or {'start_date': today.replace(day=1), 'end_date': today})
    results = []
    warehouse = None
    
    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        end_date_inclusive = end_date + timedelta(days=1)
        warehouse = form.cleaned_data['warehouse']
        
        qs = Transaction.objects.all()
        if warehouse: qs = qs.filter(warehouse=warehouse)
        elif not request.user.is_staff:
            user_whs = get_user_warehouses(request.user)
            qs = qs.filter(warehouse__in=user_whs)
            
        # Сортуємо спочатку по категорії, потім по назві
        materials = Material.objects.all().select_related('category').order_by('category__name', 'name')
        
        for mat in materials:
            mat_qs = qs.filter(material=mat)
            
            start_bal = mat_qs.filter(created_at__lt=start_date).aggregate(
                total=Sum(Case(
                    When(transaction_type='IN', then=F('quantity')),
                    When(transaction_type__in=['OUT', 'TRANSFER', 'LOSS'], then=0 - F('quantity')),
                    default=0, output_field=DecimalField()
                )))['total'] or 0
            
            period_qs = mat_qs.filter(created_at__gte=start_date, created_at__lt=end_date_inclusive)
            income = period_qs.filter(transaction_type='IN').aggregate(s=Sum('quantity'))['s'] or 0
            outcome = period_qs.filter(transaction_type__in=['OUT', 'TRANSFER', 'LOSS']).aggregate(s=Sum('quantity'))['s'] or 0
            
            end_bal = start_bal + income - outcome
            
            if start_bal != 0 or income != 0 or outcome != 0:
                cat_name = mat.category.name if mat.category else "Інше"
                results.append({
                    'material': mat, 
                    'category': cat_name, # Додали категорію
                    'start_balance': start_bal, 
                    'income': income, 
                    'outcome': outcome, 
                    'end_balance': end_bal,
                    'total_value': round(end_bal * mat.current_avg_price, 2)
                })
                
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Оборотна"
        # Додали колонку "Категорія" в Excel
        ws.append(['Категорія', 'Матеріал', 'Од.', 'Початок', 'Прихід', 'Витрата', 'Кінець', 'Сума'])
        
        for row in results: 
            ws.append([
                row['category'],
                row['material'].name, 
                row['material'].unit, 
                row['start_balance'], 
                row['income'], 
                row['outcome'], 
                row['end_balance'], 
                row['total_value']
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=report.xlsx'; wb.save(response); return response
        
    return render(request, 'warehouse/period_report.html', {'form': form, 'results': results, 'warehouse': warehouse})

@login_required
@group_required('Finance')
def procurement_journal(request):
    items = OrderItem.objects.exclude(order__status__in=['draft', 'rejected', 'new']).select_related('order', 'material', 'order__warehouse', 'order__created_by', 'order__selected_supplier')
    
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    if date_from: items = items.filter(order__updated_at__date__gte=date_from)
    if date_to: items = items.filter(order__updated_at__date__lte=date_to)
    
    journal_data = []
    total_period_sum = 0
    
    for item in items:
        order_price = getattr(item.order, 'supplier_price', 0)
        price = item.supplier_price or order_price or item.material.current_avg_price
        
        qty = item.quantity_fact or item.quantity
        total_cost = round(qty * price, 2)
        total_period_sum += total_cost
        
        supplier_name = "-"
        if item.order.selected_supplier: supplier_name = item.order.selected_supplier.name
        elif item.order.supplier_info: supplier_name = item.order.supplier_info
        
        journal_data.append({
            'order_id': item.order.id, 
            'created_at': item.order.created_at, 
            'status': item.order.get_status_display(),
            'manager': item.order.created_by.get_full_name() or item.order.created_by.username if item.order.created_by else "-",
            'warehouse': item.order.warehouse.name, 
            'material': item.material.name, 
            'quantity': qty, 
            'unit': item.material.unit,
            'supplier': supplier_name,
            'price_unit': price, 
            'total_cost': total_cost, 
            'note': item.order.note or "-", 
            'purchase_date': item.order.updated_at
        })
        
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(['ID', 'Дата', 'Матеріал', 'К-сть', 'Сума', 'Постачальник'])
        for r in journal_data:
            dt_created = make_naive(r['created_at']) if r['created_at'] else ""
            ws.append([r['order_id'], dt_created, r['material'], r['quantity'], r['total_cost'], r['supplier']])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=procurement.xlsx'; wb.save(response); return response
        
    return render(request, 'warehouse/procurement_journal.html', {'journal': journal_data, 'total_period_sum': total_period_sum, 'date_from': date_from, 'date_to': date_to})

@login_required
@group_required('Finance')
def financial_report(request):
    base_orders = Order.objects.exclude(status__in=['draft', 'rejected', 'new']).select_related('warehouse', 'selected_supplier').prefetch_related('items__material')
    
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    if date_from: base_orders = base_orders.filter(updated_at__date__gte=date_from)
    if date_to: base_orders = base_orders.filter(updated_at__date__lte=date_to)
    
    orders_data, total_spent = [], 0
    
    for order in base_orders:
        order_cost = 0
        order_price = getattr(order, 'supplier_price', 0)
        for item in order.items.all():
            price = item.supplier_price or order_price or item.material.current_avg_price
            qty = item.quantity_fact or item.quantity
            order_cost += float(qty) * float(price)
            
        total_spent += order_cost
        orders_data.append({
            'order': order, 
            'cost': order_cost, 
            'warehouse_name': order.warehouse.name, 
            'supplier_name': order.selected_supplier.name if order.selected_supplier else (order.supplier_info or "Інше")
        })
        
    warehouses_stats = {}
    for item in orders_data:
        wh = item['warehouse_name']
        if wh not in warehouses_stats:
            limit = Warehouse.objects.filter(name=wh).first().budget_limit
            warehouses_stats[wh] = {'spent': 0, 'limit': float(limit), 'overrun': 0}
        warehouses_stats[wh]['spent'] += item['cost']
        
    for wh, data in warehouses_stats.items():
        if data['spent'] > data['limit']: data['overrun'] = data['spent'] - data['limit']
        
    suppliers_stats = {}
    for item in orders_data:
        sup = item['supplier_name']
        suppliers_stats[sup] = suppliers_stats.get(sup, 0) + item['cost']
    sorted_suppliers = sorted(suppliers_stats.items(), key=lambda x: x[1], reverse=True)[:5]
    
    dates_stats = {}
    for item in orders_data:
        key = item['order'].updated_at.strftime('%Y-%m')
        dates_stats[key] = dates_stats.get(key, 0) + item['cost']
    sorted_dates = sorted(dates_stats.items())
    
    return render(request, 'warehouse/financial_report.html', {
        'total_spent': round(total_spent, 2), 'warehouses_stats': warehouses_stats,
        'wh_labels': json.dumps(list(warehouses_stats.keys())), 'wh_spent': json.dumps([round(d['spent'],2) for d in warehouses_stats.values()]), 'wh_limits': json.dumps([round(d['limit'],2) for d in warehouses_stats.values()]),
        'sup_labels': json.dumps([x[0] for x in sorted_suppliers]), 'sup_data': json.dumps([round(x[1],2) for x in sorted_suppliers]),
        'date_labels': json.dumps([x[0] for x in sorted_dates]), 'date_values': json.dumps([round(x[1],2) for x in sorted_dates]),
        'date_from': date_from, 'date_to': date_to
    })

@login_required
def planning_report(request):
    if not request.user.is_staff: return redirect('index')
    items = OrderItem.objects.exclude(order__status__in=['draft', 'completed', 'rejected']).select_related('order', 'material', 'order__warehouse', 'order__created_by').order_by('order__expected_date')
    
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    if date_from: items = items.filter(order__expected_date__gte=date_from)
    if date_to: items = items.filter(order__expected_date__lte=date_to)
    
    report_data = []
    for item in items:
        status_label = "На погодженні" if item.order.status in ['new', 'rfq'] else ("На закупівлі" if item.order.status in ['purchasing', 'approved'] else "В дорозі")
        status_color = "warning" if item.order.status in ['new', 'rfq'] else ("info" if item.order.status in ['purchasing', 'approved'] else "primary")
        
        report_data.append({
            'order': item.order, 
            'warehouse': item.order.warehouse.name, 
            'material': item.material.name, 
            'quantity': item.quantity, 
            'unit': item.material.unit, 
            'expected_date': item.order.expected_date, 
            'responsible': item.order.created_by.username if item.order.created_by else "-", 
            'status_label': status_label, 
            'status_color': status_color
        })
        
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(['Об\'єкт', 'Матеріал', 'К-сть', 'Стан'])
        for r in report_data: 
            ws.append([r['warehouse'], r['material'], r['quantity'], r['status_label']])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=plan.xlsx'; wb.save(response); return response
        
    return render(request, 'warehouse/planning_report.html', {'report_data': report_data, 'date_from': date_from, 'date_to': date_to})

@login_required
def stock_balance_report(request):
    if not request.user.is_staff: return redirect('index')
    warehouses, materials = Warehouse.objects.all(), Material.objects.all()
    wh_id, mat_id = request.GET.get('warehouse'), request.GET.get('material')
    
    stock_data, total_value = [], 0
    targets = warehouses.filter(id=wh_id) if wh_id else warehouses
    
    for wh in targets:
        for item in get_warehouse_balance(wh, mat_id):
            status = 'critical' if item['min_limit'] > 0 and item['quantity'] <= item['min_limit'] else 'ok'
            stock_data.append({
                'warehouse': wh.name, 'material': item['name'], 'material_id': item['id'],
                'unit': item['unit'], 'quantity': item['quantity'], 'min_limit': item['min_limit'],
                'avg_price': item['price'], 'total_value': item['total_sum'], 'status': status
            })
            total_value += item['total_sum']
            
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(['Склад', 'Матеріал', 'Залишок', 'Сума'])
        for r in stock_data: ws.append([r['warehouse'], r['material'], r['quantity'], r['total_value']])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=balance.xlsx'; wb.save(response); return response
        
    return render(request, 'warehouse/stock_balance_report.html', {
        'stock_data': stock_data, 'total_value': total_value, 'warehouses': warehouses, 'materials': materials, 
        'selected_wh': int(wh_id) if wh_id else None, 'selected_mat': int(mat_id) if mat_id else None
    })

@login_required
def movement_history(request):
    if not request.user.is_staff: return redirect('index')
    transactions = Transaction.objects.all().select_related('material', 'warehouse', 'order', 'created_by').order_by('-created_at')
    
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    op_type, mat_id, wh_id = request.GET.get('type'), request.GET.get('material'), request.GET.get('warehouse')
    
    if date_from: transactions = transactions.filter(created_at__date__gte=date_from)
    if date_to: transactions = transactions.filter(created_at__date__lte=date_to)
    if op_type: transactions = transactions.filter(transaction_type=op_type)
    if mat_id: transactions = transactions.filter(material_id=mat_id)
    if wh_id: transactions = transactions.filter(warehouse_id=wh_id)
    
    history_data = []
    for t in transactions:
        sender, receiver = "-", "-"
        if t.transaction_type == 'IN':
            if t.order:
                if t.order.selected_supplier: sender = t.order.selected_supplier.name
                elif t.order.source_warehouse: sender = t.order.source_warehouse.name
                else: sender = t.order.supplier_info or "Зовнішній"
            else: sender = "Коригування"
            receiver = t.warehouse.name
        elif t.transaction_type == 'OUT':
            sender = t.warehouse.name
            receiver = "Будівництво"
        elif t.transaction_type == 'TRANSFER':
            sender = t.warehouse.name
            receiver = t.order.warehouse.name if t.order else "Склад"
        elif t.transaction_type == 'LOSS':
            sender = t.warehouse.name
            receiver = "Списання"
        
        user_display = "-"
        if t.created_by: user_display = t.created_by.get_full_name() or t.created_by.username
        elif t.order and t.order.created_by: user_display = t.order.created_by.get_full_name() or t.order.created_by.username

        history_data.append({
            'date': t.created_at, 'type_code': t.transaction_type, 'type_label': t.get_transaction_type_display(),
            'material': t.material.name, 'material_id': t.material.id,
            'quantity': t.quantity, 'unit': t.material.unit, 'sender': sender, 'receiver': receiver,
            'user': user_display,
            'basis': f"Заявка #{t.order.id}" if t.order else t.description, 'order_id': t.order.id if t.order else None, 'desc': t.description
        })
        
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(['Дата', 'Тип', 'Матеріал', 'К-сть', 'Звідки/Куди'])
        for r in history_data:
            dt = r['date']
            if dt: dt = make_naive(dt)
            ws.append([dt, r['type_label'], r['material'], r['quantity'], f"{r['sender']} -> {r['receiver']}"])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=history.xlsx'; wb.save(response); return response
        
    return render(request, 'warehouse/movement_history.html', {
        'history': history_data, 'materials': Material.objects.all(), 'warehouses': Warehouse.objects.all(),
        'f_date_from': date_from, 'f_date_to': date_to, 'f_type': op_type, 'f_mat': int(mat_id) if mat_id else None, 'f_wh': int(wh_id) if wh_id else None
    })

@login_required
def writeoff_report(request):
    if not request.user.is_staff: return redirect('index')
    transactions = Transaction.objects.filter(transaction_type__in=['OUT', 'LOSS']).order_by('-created_at')
    wh_id, mat_id, reason = request.GET.get('warehouse'), request.GET.get('material'), request.GET.get('reason')
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    if date_from: transactions = transactions.filter(created_at__date__gte=date_from)
    if date_to: transactions = transactions.filter(created_at__date__lte=date_to)
    if wh_id: transactions = transactions.filter(warehouse_id=wh_id)
    if mat_id: transactions = transactions.filter(material_id=mat_id)
    if reason: transactions = transactions.filter(transaction_type=reason)
    stats = {'work_sum': 0, 'loss_sum': 0, 'total_sum': 0}
    data = []
    for t in transactions:
        sum_val = t.quantity * t.price
        stats['total_sum'] += sum_val
        if t.transaction_type == 'OUT': stats['work_sum'] += sum_val
        else: stats['loss_sum'] += sum_val
        photo_url = None
        if t.photo: photo_url = t.photo.url
        elif t.order and t.order.proof_photo: photo_url = t.order.proof_photo.url
        
        author_display = "-"
        if t.created_by: author_display = t.created_by.get_full_name() or t.created_by.username
        elif t.order and t.order.created_by: author_display = t.order.created_by.get_full_name() or t.order.created_by.username

        data.append({
            'date': t.created_at, 'warehouse': t.warehouse.name, 'material': t.material.name, 'material_id': t.material.id,
            'quantity': t.quantity, 'unit': t.material.unit, 'type_code': t.transaction_type, 'type_display': t.get_transaction_type_display(),
            'author': author_display,
            'reason': t.description,
            'photo': photo_url, 'sum': round(sum_val, 2)
        })
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(['Дата', 'Тип', 'Матеріал', 'Сума', 'Причина'])
        for r in data:
            dt = r['date']
            if dt: dt = make_naive(dt)
            ws.append([dt, r['type_code'], r['material'], r['sum'], r['reason']])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=writeoff.xlsx'; wb.save(response); return response
    return render(request, 'warehouse/writeoff_report.html', {
        'report_data': data, 'stats': stats, 'warehouses': Warehouse.objects.all(), 'materials': Material.objects.all(),
        'f_wh': int(wh_id) if wh_id else None, 'f_mat': int(mat_id) if mat_id else None, 'f_reason': reason,
        'f_date_from': date_from, 'f_date_to': date_to
    })

@login_required
def transfer_journal(request):
    if not request.user.is_staff: return redirect('index')
    items = OrderItem.objects.filter(order__source_warehouse__isnull=False).exclude(order__status='draft').select_related('order', 'material', 'order__warehouse', 'order__source_warehouse').order_by('-order__created_at')
    date_from, date_to, status = request.GET.get('date_from'), request.GET.get('date_to'), request.GET.get('status')
    if date_from: items = items.filter(order__updated_at__date__gte=date_from)
    if date_to: items = items.filter(order__updated_at__date__lte=date_to)
    if status: items = items.filter(order__status=status)
    journal_data = []
    for item in items:
        st_label = "В дорозі" if item.order.status == 'in_transit' else ("Доставлено" if item.order.status == 'completed' else "Готується")
        st_class = "primary" if item.order.status == 'in_transit' else ("success" if item.order.status == 'completed' else "warning")
        journal_data.append({
            'date': item.order.created_at, 'material': item.material.name, 'quantity': item.quantity, 'unit': item.material.unit,
            'source': item.order.source_warehouse.name, 'destination': item.order.warehouse.name, 
            'initiator': item.order.created_by.username if item.order.created_by else "-",
            'status_label': st_label, 'status_class': st_class, 'id': item.order.id
        })
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(['Дата', 'Звідки', 'Куди', 'Матеріал'])
        for r in journal_data:
            dt = r['date']
            if dt: dt = make_naive(dt)
            ws.append([dt, r['source'], r['destination'], r['material']])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=transfers.xlsx'; wb.save(response); return response
    return render(request, 'warehouse/transfer_journal.html', {'journal': journal_data, 'f_date_from': date_from, 'f_date_to': date_to, 'f_status': status})

@login_required
def transfer_analytics(request):
    if not request.user.is_staff: return redirect('index')
    transfers = Order.objects.filter(source_warehouse__isnull=False).exclude(status__in=['draft', 'rejected'])
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    if date_from: transfers = transfers.filter(updated_at__date__gte=date_from)
    if date_to: transfers = transfers.filter(updated_at__date__lte=date_to)
    total_moves = transfers.count()
    top_mat_qs = OrderItem.objects.filter(order__in=transfers).values('material__name').annotate(c=Count('id')).order_by('-c')[:5]
    routes = transfers.values('source_warehouse__name', 'warehouse__name').annotate(trips=Count('id')).order_by('-trips')[:10]
    completed = transfers.filter(status__in=['completed', 'approved'])
    total_sec = sum([(t.updated_at - t.created_at).total_seconds() for t in completed])
    avg_hours = round((total_sec / completed.count()) / 3600, 1) if completed.exists() else 0
    loss_cases = OrderItem.objects.filter(order__in=completed, quantity_fact__lt=F('quantity')).count()
    return render(request, 'warehouse/transfer_analytics.html', {
        'total_moves': total_moves, 'avg_hours': avg_hours, 'loss_cases': loss_cases,
        'mat_labels': json.dumps([m['material__name'] for m in top_mat_qs]), 'mat_data': json.dumps([m['c'] for m in top_mat_qs]),
        'route_labels': json.dumps([f"{r['source_warehouse__name']}->{r['warehouse__name']}" for r in routes]), 'route_data': json.dumps([r['trips'] for r in routes]),
        'routes': routes, 'date_from': date_from, 'date_to': date_to
    })

@login_required
@group_required('Finance')
def objects_comparison(request):
    warehouses = Warehouse.objects.all()
    data, labels, budget_d, spent_d = [], [], [], []
    for wh in warehouses:
        budget = wh.budget_limit
        spent = Transaction.objects.filter(warehouse=wh, transaction_type='OUT').aggregate(t=Sum(F('quantity')*F('price'), output_field=DecimalField()))['t'] or 0
        balance = budget - spent
        reserved = OrderItem.objects.filter(order__warehouse=wh, order__status__in=['new', 'approved', 'purchasing']).aggregate(t=Sum(F('quantity') * F('material__current_avg_price'), output_field=DecimalField()))['t'] or 0
        status = 'ok'
        if balance < 0: status = 'critical'
        elif balance < (budget * Decimal('0.2')): status = 'warning'
        data.append({'name': wh.name, 'responsible': wh.responsible.username if wh.responsible else "-", 'budget': budget, 'spent': spent, 'balance': balance, 'active_commitments': reserved, 'forecast': balance - reserved, 'status': status})
        labels.append(wh.name); budget_d.append(float(budget)); spent_d.append(float(spent))
    data.sort(key=lambda x: x['balance'])
    return render(request, 'warehouse/objects_comparison.html', {'report_data': data, 'chart_labels': json.dumps(labels), 'chart_budget': json.dumps(budget_d), 'chart_spent': json.dumps(spent_d)})

@login_required
@group_required('Finance')
def suppliers_rating(request):
    suppliers = Supplier.objects.all()
    data = []
    date_from, date_to = request.GET.get('date_from'), request.GET.get('date_to')
    for s in suppliers:
        orders = Order.objects.filter(selected_supplier=s).exclude(status='draft')
        if date_from: orders = orders.filter(created_at__date__gte=date_from)
        if date_to: orders = orders.filter(created_at__date__lte=date_to)
        if not orders.exists(): continue
        spent = Transaction.objects.filter(order__in=orders, transaction_type='IN').aggregate(s=Sum(F('quantity')*F('price')))['s'] or 0
        problems = 0
        if orders.filter(status='rejected').exists(): problems += orders.filter(status='rejected').count()
        items_with_issues = OrderItem.objects.filter(order__in=orders, quantity_fact__lt=F('quantity')*Decimal('0.99')).count()
        problems += items_with_issues
        reliability = round(100 - (problems / orders.count() * 100), 1)
        if reliability < 0: reliability = 0
        data.append({
            'supplier': s, 'total_spent': spent, 'orders_count': orders.count(),
            'avg_delivery': 0, 'reliability': reliability, 'rel_class': 'success' if reliability > 80 else 'warning',
            'materials_list': ", ".join([m.name for m in s.materials.all()[:3]])
        })
    data.sort(key=lambda x: x['total_spent'], reverse=True)
    return render(request, 'warehouse/suppliers_rating.html', {'rating_data': data, 'date_from': date_from, 'date_to': date_to})

@login_required
def problem_areas(request):
    if not request.user.is_staff: return redirect('index')
    shortages = []
    completed_items = OrderItem.objects.filter(order__status='completed', quantity_fact__lt=F('quantity')*Decimal('0.99'))
    for item in completed_items:
        fact = item.quantity_fact or 0
        diff = item.quantity - fact
        oprice = getattr(item.order, 'supplier_price', 0)
        price = item.supplier_price or oprice or item.material.current_avg_price
        shortages.append({'order': item.order, 'material': item.material, 'diff': diff, 'sum': round(diff * price, 2)})
    overdue_items = OrderItem.objects.filter(order__status__in=['approved', 'purchasing', 'in_transit'], order__expected_date__lt=timezone.now().date())
    critical_wh = []
    for wh in Warehouse.objects.all():
        spent = Transaction.objects.filter(warehouse=wh, transaction_type='OUT').aggregate(t=Sum(F('quantity')*F('price'), output_field=DecimalField()))['t'] or 0
        if spent > wh.budget_limit:
            top_expenses = Transaction.objects.filter(warehouse=wh, transaction_type='OUT').values('material__name').annotate(total=Sum(F('quantity')*F('price'))).order_by('-total')[:3]
            critical_wh.append({'name': wh.name, 'limit': wh.budget_limit, 'overrun': spent - wh.budget_limit, 'percent': round(((spent - wh.budget_limit)/wh.budget_limit)*100, 1), 'top_expenses': top_expenses})
    losses = Transaction.objects.filter(transaction_type='LOSS').order_by('-created_at')[:20]
    loss_stats = Transaction.objects.filter(transaction_type='LOSS').values('warehouse__name').annotate(total=Sum(F('quantity')*F('price'), output_field=DecimalField()), count=Count('id')).order_by('-total')[:5]
    return render(request, 'warehouse/problem_areas.html', {
        'shortages': shortages, 'overdue': overdue_items, 'critical_wh': critical_wh, 'recent_losses': losses, 'loss_by_wh': loss_stats, 'today': timezone.now().date()
    })

@login_required
@group_required('Finance')
def savings_report(request):
    date_from = request.GET.get('date_from'); date_to = request.GET.get('date_to')
    transactions = Transaction.objects.filter(transaction_type='OUT').select_related('material', 'warehouse')
    if date_from: transactions = transactions.filter(created_at__date__gte=date_from)
    if date_to: transactions = transactions.filter(created_at__date__lte=date_to)
    report_data = []
    total_savings = 0
    for t in transactions:
        market_price = t.material.market_price
        if market_price == 0: market_price = t.price 
        diff_per_unit = market_price - t.price
        total_saving_on_transaction = diff_per_unit * t.quantity
        if total_saving_on_transaction != 0:
            total_savings += total_saving_on_transaction
            report_data.append({'date': t.created_at, 'warehouse': t.warehouse.name, 'material': t.material.name, 'qty': t.quantity, 'unit': t.material.unit, 'cost_price': t.price, 'market_price': market_price, 'diff': diff_per_unit, 'saving': total_saving_on_transaction})
    report_data.sort(key=lambda x: x['saving'], reverse=True)
    return render(request, 'warehouse/savings_report.html', {'report_data': report_data, 'total_savings': round(total_savings, 2), 'date_from': date_from, 'date_to': date_to})

@login_required
def export_stock_report(request):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Залишки"
    ws.append(['Склад', 'Матеріал', 'Кількість', 'Од.', 'Сума'])
    for wh in get_user_warehouses(request.user):
        for item in get_warehouse_balance(wh):
            ws.append([wh.name, item['name'], item['quantity'], item['unit'], item['total_sum']])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'); response['Content-Disposition'] = 'attachment; filename=stock.xlsx'; wb.save(response); return response

@login_required
def global_audit_log(request):
    if not request.user.is_staff: return redirect('index')
    logs = AuditLog.objects.all().select_related('user').order_by('-timestamp')
    f_user = request.GET.get('user'); f_action = request.GET.get('action'); f_date = request.GET.get('date')
    if f_user: logs = logs.filter(user__username__icontains=f_user)
    if f_action: logs = logs.filter(action_type=f_action)
    if f_date: logs = logs.filter(timestamp__date=f_date)
    logs = logs[:200]
    return render(request, 'warehouse/audit_log.html', {'logs': logs, 'action_types': AuditLog.ACTION_TYPES})